from typing import Type
from PySide2.QtCore import QDate
from PySide2.QtWidgets import QApplication,QMessageBox,QMainWindow
from PySide2.QtGui import QStandardItemModel,QStandardItem
from weibao_ui import Ui_MainWindow
import pyodbc,datetime,openpyxl,os
from openpyxl.styles import Font,PatternFill,Border,Side,Alignment
from openpyxl.chart import Reference,BarChart,Series,BarChart3D,PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText

cur_year = datetime.datetime.now().year
cur_month = datetime.datetime.now().month
cur_day = datetime.datetime.now().day
cur_date = str(cur_year) + "-" + ("00" + str(cur_month))[-2:] + "-" +("00" + str(cur_day))[-2:]

class weibao(Ui_MainWindow,QMainWindow):
    def __init__(self,parent=None):
        super(Ui_MainWindow,self).__init__(parent=None)
        self.setupUi(self)
        self.connect_db()
        self.Dis_Data()
        self.Pb_Query.clicked.connect(self.Dis_Data)
        self.Cb_State.addItems(['','使用中','闲置中','已报废'])
        self.Cb_baoxiu.addItems(['','一年','二年','三年','四年','五年'])
        self.Cb_Level.addItems(['','高端电脑','中端电脑','低端电脑'])
        self.Cb_Memory.addItems(['','2G','4G','6G','8G','16G','32G','其它'])
        self.Cb_SolidDisk.addItems(['','120G','240G','480G','其它'])
        self.Cb_MachineDisk.addItems(['','320G','500G','1000G','2000G'])
        self.Cb_Branch.addItems(['','戴尔','华为','联想','组装机','其它'])
        self.Cb_Type.addItems(['','笔记本电脑','品牌台式机','组装台式机','打印机','考勤机','服务器','其它设备'])
        self.Cb_CPU.addItems(['','I3','I5','I7','其它'])
        self.Cb_Provider.addItems(['','京东','天猫'])
        self.comboBox_operate_user.addItems(['','周松辉','陈一迪','杨元政'])
        self.comboBox_wb_type.addItems(['','硬件维护','软件维护'])
        self.comboBox_ios.addItems(['','Windos10','Windows8','Windows7','Windows Xp','Windows2003','其它'])
        self.comboBox_State.addItems(['','使用中','闲置中','已报废'])

        #tableview单击操作
        self.Tb_Device.clicked.connect(self.Set_Value)
        self.Pb_Update.clicked.connect(self.Upd_value)
        self.Tb_WB.clicked.connect(self.Set_wb_info)
        self.pushButton_analysis.clicked.connect(self.mac_analysis)
        self.pushButton_upd_weibao.clicked.connect(self.Update_Wb_Info)
        self.pushButton_new_wb.clicked.connect(self.new_wb_info)
        #self.pushButton_wb_record_del.clicked.connect(self.wb_record_del)
        self.pushButton_Trace.clicked.connect(self.Dev_trace)
    def connect_db(self):
        server = '192.168.1.254'
        database = 'ZWERP'
        username = 'sa'
        password = 'zhaowenerp.369'
        conn = self.cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        self.cursor = conn.cursor()
        

    def Dis_Data(self):
        device_id = self.Le_ID.text()
        if len(device_id.strip())==0:
            device_id = '%'
        else:
            device_id = '%'+device_id+'%'
        user_info = self.Le_User.text()
        if len(user_info.strip())==0:
            user_info = '%'
        else:
            user_info = '%'+user_info+'%'
        ip_addr = self.lineEdit_IP_Query.text()
        if len(ip_addr.strip())==0:
            ip_addr = '%'
        else:
            ip_addr = '%'+ip_addr+'%'
        state = self.comboBox_State.currentText()
        if len(state.strip())==0:
            state = '%'
        else:
            state = '%'+state+'%'
        
        parames = ([device_id,user_info,ip_addr,state])
        sql = 'SELECT deviceid,rtrim(ltrim(a.supplier)),rtrim(ltrim(a.model)),a.dname,rtrim(ltrim(a.level)),rtrim(ltrim(a.position)),rtrim(ltrim(a.cpu)),rtrim(ltrim(a.mem)),a.solid_disk,rtrim(ltrim(a.harddisk)),rtrim(ltrim(a.state)), \
            a.ios,a.ip_addr,rtrim(ltrim(a.description)) from d_base a where a.deviceid  like ? and a.position like ? and a.ip_addr like ? and a.state like ?  order by a.deviceid'
        result = self.cursor.execute(sql,parames).fetchall()
        model = QStandardItemModel(len(result),2)
        model.setHorizontalHeaderLabels(['硬件编码','供应商','型号','类型','级别','使用人','CPU','内存','固态硬盘','机械硬盘','状态','操作系统','IP地址','备注'])
        for row in range(len(result)):
            for column in range(12):
                item = QStandardItem(result[row][column])
                model.setItem(row,column,item)
        self.Tb_Device.setModel(model)
        self.statusBar.showMessage('共查询到'+str(len(result))+'笔数据!')
        self.Tb_Device.selectRow(0)
        self.Set_Value()
        self.Dis_wb_record()
        self.Set_wb_info()


    def Set_Value(self):
        
        #获取选择行的内容
        index = self.Tb_Device.currentIndex()
        self.pc_id = self.Tb_Device.model().index(index.row(),0).data()
        #执行查找出来的数据，设置到相应的栏位中
        parames = ([self.pc_id])
        sql = 'select rtrim(ltrim(a.supplier)),rtrim(ltrim(a.model)),rtrim(ltrim(a.level)),rtrim(ltrim(a.position)),rtrim(ltrim(a.cpu)),rtrim(ltrim(a.mem)),rtrim(ltrim(a.harddisk)),rtrim(ltrim(a.state)),a.solid_disk,rtrim(ltrim(a.description)), \
            a.ios,a.ip_addr,a.dname from d_base a  where a.deviceid =? '
        result = self.cursor.execute(sql,parames).fetchone()
        if result is not None:
            supplier = result[0]
            model = result[1]
            level = result[2]
            nm = result[3]
            cpu = result[4]  
            mem = result[5]
            harddisk = result[6]
            state = result[7]
            solid_disk = result[8]
            description = result[9]
            ios = result[10]
            ip_addr = result[11]
            dname = result[12]
            self.Cb_Branch.setCurrentText(supplier)
            self.Cb_Level.setCurrentText(level)
            self.Le_Userinfo.setText(nm)
            self.Cb_CPU.setCurrentText(cpu)
            self.Cb_Memory.setCurrentText(mem)
            self.Cb_MachineDisk.setCurrentText(harddisk)
            self.Cb_State.setCurrentText(state)
            self.Cb_SolidDisk.setCurrentText(solid_disk)                            
            self.lineEdit_Remark.setText(description)
            self.comboBox_ios.setCurrentText(ios)
            self.lineEdit_Ip.setText(ip_addr)
            self.lineEdit_model.setText(dname)
            #执行变更程序
            
            self.Dis_wb_record()    
            self.Show_ch_Record()                             
        else:
            supplier = ''
            model = ''
            level = ''
            nm = ''
            cpu = ''  
            mem = ''
            harddisk = ''
            state = ''
            solid_disk = ''
            description = ''
            ios = ''
            ip_addr = ''
            dname = ''
            self.Cb_Branch.setCurrentText(supplier)
            self.Cb_Level.setCurrentText(level)
            self.Le_Userinfo.setText(nm)
            self.Cb_CPU.setCurrentText(cpu)
            self.Cb_Memory.setCurrentText(mem)
            self.Cb_MachineDisk.setCurrentText(harddisk)
            self.Cb_State.setCurrentText(state)
            self.Cb_SolidDisk.setCurrentText(solid_disk)                            
            self.lineEdit_Remark.setText(description)
            self.comboBox_ios.setCurrentText(ios)
            self.lineEdit_Ip.setText(ip_addr)
            self.lineEdit_model.setText(dname)


        #名称匹配不精确，单独列出来增加栏位的准确性
        parames1 = ([self.pc_id,'在职'])
        sql1 = 'select b.position,c.deptnm from d_base a left join employee b on a.position = b.name left join dept c on b.deptid = c.deptid  where a.deviceid = ? and b.state=?'
        result1 = self.cursor.execute(sql1,parames1).fetchone()
        if result1 is not None:
            position = result1[0]
            deptnm = result1[1]
            self.lineEdit_Dept.setText(deptnm)
            self.Le_Title.setText(position)
        else:
            position =''
            deptnm = ''
            self.lineEdit_Dept.setText("")
            self.Le_Title.setText("")

        #在设置设备参数值时,保存为全局变量。在设备变更需求时可以用上
        #self.old_supplier = supplier
        #self.old_model = model
        #self.old_level = level
        self.old_nm = nm
        #self.old_cpu = cpu
        #self.old_mem =  mem
        #self.old_harddisk = harddisk
        #self.old_state = state
        #self.old_solid_disk = solid_disk
        self.old_description = description
        #self.old_ios = ios
        #self.old_nm = nm
        
        


    def Upd_value(self):
        supplier = self.Cb_Branch.currentText()
        level = self.Cb_Level.currentText()
        position = self.Le_Userinfo.text()

        #异常判断,若输入了不存在的改名则变更失败。若有重名的,则进行提示
        sql = 'select count(name) from employee where name =?'
        counter = self.cursor.execute(sql,position).fetchone()
        if counter[0] < 1 and len(position) != 0:
            self.statusBar.showMessage('姓名输入错误,变更失败!')
            return
        if counter[0] > 1:
            self.statusBar.showMessage('存在同名人员,请确认组别职称是否正确!')

        cpu = self.Cb_CPU.currentText()
        mem = self.Cb_Memory.currentText()
        harddisk = self.Cb_MachineDisk.currentText()
        solid_disk = self.Cb_SolidDisk.currentText()
        state = self.Cb_State.currentText()
        type = self.Cb_Type.currentText()
        ip_addr = self.lineEdit_Ip.text()
        ios = self.comboBox_ios.currentText()
        remark = self.lineEdit_Remark.text()
        parames = ([supplier,level,position,cpu,mem,harddisk,solid_disk,state,type,ip_addr,ios,remark,self.pc_id])
        sql = 'update d_base set supplier = ?,level=?,position=?,cpu=?,mem=?,harddisk=?,solid_disk=?,state=?,device_type=?,ip_addr=?,ios=?,description=? where deviceid=?'
        counter = self.cursor.execute(sql,parames).rowcount
        if counter > 0:
            self.statusBar.showMessage('操作成功，总共更新了'+str(counter)+'条数据.')
            self.cursor.commit()
            self.Dis_wb_record()
            #self.Dis_Data()
        else:
            self.statusBar.showMessage('更新失败！')
            return

        #将更新记录保存到变更记录中
        #针对于使用者跟备注的变更作了记录更新,若后续有其它的需求，增加相应的柆位就行
        if self.old_nm != position or self.old_description != remark:
            #生成变更序列号
            sql_series = '''select 'CH' + '0000' + convert(varchar,isnull(right(max(ch_id),4),0)+1) from ch_record'''
            result = self.cursor.execute(sql_series).fetchone()
            series = result[0]
            if len(position)==0:
                position = '闲置'
            if self.old_nm != position:
                remark_a = self.old_nm + '-->' +  position + ',' +self.old_description + '--->' + remark
            else:
                remark_a = self.old_description + '--->' + remark

            parames = ([self.pc_id,cur_date,self.old_nm,position,series,remark_a])
            sql = 'insert into ch_record(deviceid,ch_date,user_old,user_new,ch_id,remark) values (?,?,?,?,?,?)'
            self.cursor.execute(sql,parames)
            self.cursor.commit()
        #重新检索变更记录
        self.Set_Value()
        self.Show_ch_Record()


        #更新完后需要实时更新旧的变量值
        self.old_nm = position
        self.old_description = remark
        
        
    def Dis_wb_record(self):
        sql = 'select wb_id,wb_date from wb_record where deviceid =? '
        result = self.cursor.execute(sql,self.pc_id).fetchall()
        model = QStandardItemModel(len(result),2)
        model.setHorizontalHeaderLabels(['维保编号','日期'])
        for row in range(len(result)):
            for column in range(2):
                item = QStandardItem(result[row][column])
                model.setItem(row,column,item)
        self.Tb_WB.setModel(model)
        self.Tb_WB.selectRow(0)
        self.Set_wb_info()

    def Set_wb_info(self):
        index = self.Tb_WB.currentIndex()
        wb_id = self.Tb_WB.model().index(index.row(),0).data()
        #若为新增的维保号，则不进行操作
        sql0 = 'select count(*) from wb_record where wb_id = ?'
        counter = self.cursor.execute(sql0,wb_id).fetchone()
        if counter[0] > 0:
            sql = 'select operater,wb_type,wb_date,wb_desc,wb_solution  from wb_record where wb_id = ?'
            result = self.cursor.execute(sql,wb_id).fetchone()
            operate = result[0]
            wb_type = result[1]
            wb_date = result[2]
            wb_desc = result[3]
            wb_solution = result[4]
            self.comboBox_operate_user.setCurrentText(operate)
            self.dateEdit_wbdate.setDate(QDate(int(wb_date[0:4]),int(wb_date[5:7]),int(wb_date[-2:])))
            self.comboBox_wb_type.setCurrentText(wb_type)
            self.Te_Wb_Info.setText(wb_desc)
            self.Te_Solution.setText(wb_solution)
        else:
            self.comboBox_operate_user.setCurrentText("")
            self.comboBox_wb_type.setCurrentText((""))
            self.Te_Wb_Info.setText((""))
            self.Te_Solution.setText((""))



    def Update_Wb_Info(self):
        #若无记录，则需要先新增记录
        index = self.Tb_WB.currentIndex()
        wb_id = self.Tb_WB.model().index(index.row(),0).data()
        if wb_id is None:
            self.statusBar.showMessage('请先新增维保编号!')
            return
        wb_date = self.dateEdit_wbdate.text()
        wb_operate = self.comboBox_operate_user.currentText()
        wb_type = self.comboBox_wb_type.currentText()
        wb_desc = self.Te_Wb_Info.toPlainText()
        wb_solution = self.Te_Solution.toPlainText()
        sql = "update wb_record set operater = ?,wb_type=?,wb_date=?,wb_desc=?,wb_solution=? where wb_id = ?"
        parames = ([wb_operate,wb_type,wb_date,wb_desc,wb_solution,wb_id])
        self.cursor.execute(sql,parames)
        counter = self.cursor.execute(sql,parames).rowcount
        if counter > 0:
            self.statusBar.showMessage('操作成功，总共更新了'+str(counter)+'条数据.')
            self.cursor.commit()
            self.Set_wb_info()
        else:
            self.statusBar().showMessage('更新失败！')
        
    def new_wb_info(self):
        #是否存在焦点，若不存在则略过
        index = self.Tb_Device.currentIndex()
        self.pc_id = self.Tb_Device.model().index(index.row(),0).data()
        if len(self.pc_id.strip())==0:
            self.statusBar().showMessage('请先选择维护的主机编号!')
            return
        #生成维保编号
        sql = "SELECT 'WB' + right('00000' + convert(nvarchar,convert(int,right(MAX(WB_ID),5))+1),5) FROM wb_record"
        result = self.cursor.execute(sql).fetchone()
        wb_id = result[0]
        parames=([wb_id,self.pc_id,cur_date])
        sql2 = 'insert into wb_record(wb_id,deviceid,wb_date) values (?,?,?)'
        counter = self.cursor.execute(sql2,parames).rowcount
        if counter > 0:
            self.statusBar.showMessage('新增成功!.')
            self.cursor.commit()
        else:
            self.statusBar.showMessage('新增失败！')
        self.Dis_wb_record()
        self.dateEdit_wbdate.setDate(QDate(cur_year,cur_month,cur_day))
        

    def mac_analysis(self):

        #sql = '''select isnull(b.Deptnm,'其它组'),isnull(b.position,'无'),a.position,a.supplier,a.model,a.level,a.cpu,a.mem,a.Solid_Disk,a.harddisk,a.ios,a.state
        #from d_base a left join 
        #(select a.name,a.position,b.Deptnm from employee a left join dept b on a.deptid = b.deptid where state=?) b on a.position = b.name 
        #order by b.deptnm desc,b.position'''

        sql ='''select case  when rtrim(ltrim(a.position)) = '陈总' then '总经办' 
		     when rtrim(ltrim(a.position)) = '周总' then '总经办' 
			 when rtrim(ltrim(a.position)) = '王总' then '总经办'  
			 when rtrim(ltrim(a.position)) = '郭嘉俊' then '财务部' 
			 when len(rtrim(ltrim(a.position))) = 0 then '其它组' 
			 when isnull(b.Deptnm,'空') = '空' then '其它组' 
			 else b.Deptnm end as deptnm, 
	   isnull(b.position,''),a.position,a.supplier,a.model,a.level,a.cpu,a.mem,a.Solid_Disk,a.harddisk,a.ios,a.state from d_base a left join  
        (select a.name,a.position,b.Deptnm from employee a left join dept b on a.deptid = b.deptid where state=?) b on a.position = b.name  
        order by deptnm desc,b.position,a.position desc
           '''

        result = self.cursor.execute(sql,'在职').fetchall()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '设备使用明细'
        ws.append(['组别','职称','使用人员','品牌','型号','级别','CPU','内存','固态硬盘','机械硬盘','操作系统','设备状态'])
        for n in result:
            ws.append([n[0],n[1],n[2],n[3],n[4],n[5],n[6],n[7],n[8],n[9],n[10],n[11]])

        xfont = Font(size=14)
        xfill = PatternFill(fill_type=None,fgColor="0033CCCC")
        for column in [chr(n) for n in range(65,77)]:
            #设置字符格式
            ws[column + str(1)].font = Font(name='微软雅黑',size=14,color='00000000')
            #设置背景色
            ws[column + str(1)].fill = PatternFill("solid", fgColor="0099CCFF")
            #设置列宽
            ws.column_dimensions[column].width = 15
            #居中对齐
            ws[column + str(1)].alignment = Alignment(horizontal='center',vertical='center')
        #低端配置特殊显示    
        for row in range(3,ws.max_row):
            if ws['F'+str(row)].value is None or (ws['F'+str(row)].value).strip() == '低端' :
                for column in [chr(n) for n in range(67,77)]:
                    ws[column + str(row)].fill = PatternFill("solid", fgColor="0099CCFF")
        #闲置电脑特殊显示
        for row in range(3,ws.max_row):
            if ws['L'+str(row)].value is None or (ws['L'+str(row)].value).strip() == '闲置中' :
                for column in [chr(n) for n in range(67,77)]:
                    ws[column + str(row)].fill = PatternFill("solid", fgColor="0000FF00")
            if ws['L'+str(row)].value is None or (ws['F'+str(row)].value).strip() == '低端电脑' :                    
                for column in [chr(n) for n in range(67,77)]:
                    ws[column + str(row)].fill = PatternFill("solid", fgColor="00FF0000")
        xborder = Border(left=Side(border_style='thin',color='000000'),right=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
        for row in range(1,ws.max_row+1):
            for column in [chr(n) for n in range(65,77)]:
                ws[column + str(row)].font = Font(name='微软雅黑',size=13)
                ws[column + str(row)].border = xborder 
        #合并单元格
        start = 2
        old = ws['A2'].value
        for row in range(2,ws.max_row + 2 ):
            new = ws['A' + str(row)].value
            if new != old :
                ws.merge_cells('A' + str(start) + ":A" + str(row - 1))
                #ws['A' + str(row)].alignment = Alignment(horizontal='center',vertical='center')
                ws['A' + str(start)].alignment = Alignment(horizontal='center',vertical='center')
                start = row
                old = new
        start = 2    
        old = ws['B2'].value
        for row in range(3,ws.max_row + 2 ):
            new = ws['B' + str(row)].value
            if new != old :
                ws.merge_cells('B' + str(start) + ":B" + str(row - 1))
                ws['B' + str(row)].alignment = Alignment(horizontal='center',vertical='center')
                start = row
                old = new
        #部门机器配置柱形图
        wb.create_sheet('数据显示栏',index=1)
        wb.create_sheet('数据栏',index=2)
        #数据统一写到数据栏中
        wb.active = 2
        ws = wb.active
        ws.append(['组别','数量'])
        sql_bar = '''
        select isnull(b.Deptnm,'其它组'),count(deviceid) from d_base a left join 
        (select a.name,a.position,b.Deptnm from employee a left join dept b on a.deptid = b.deptid where state='在职') b on a.position = b.name 
        group by b.Deptnm
        '''
        result = self.cursor.execute(sql_bar).fetchall()
        for n in result:
            ws.append([n[0],n[1]])
        #数据添加完成后，转回到数据显示工作表
        max_row1 = ws.max_row
        chart1 = BarChart()
        chart1.type='col'
        chart1.style= 10
        chart1.title='部门设备明细'
        chart1.height = 21
        chart1.width = 30
        chart1.showVal=True
        data = Reference(ws,min_col=2,min_row=2,max_row=max_row1,max_col=2)
        title = Reference(ws,min_col=1,min_row=2,max_row=max_row1,max_col=1)
        chart1.add_data(data)
        chart1.set_categories(title)
        s1 = chart1.series[0]
        s1.dLbls = DataLabelList() 
        s1.dLbls.showVal = True
        s1.dLbls.showPercent = True
        #设置柱形的颜色
        #s1.graphicalProperties.line.solidFill = "00000"
        #s1.graphicalProperties.solidFill = "ff9900" 
        wb.active = 1
        ws = wb.active
        ws.add_chart(chart1,'A1')
       
        #在用与闲置圆饼图
        #返回数据栏写入汇总数据
        wb.active=2
        ws=wb.active
        #ws.append(['状态','数量'])
        sql_pie = 'select state,count(deviceid) from d_base where state <> ? group by state'
        result = self.cursor.execute(sql_pie,'已报废').fetchall()
        for n in result:
            ws.append([n[0],n[1]]) 
        max_row2 = ws.max_row
        chart2 = PieChart()
        chart2.title ='电脑使用占比图'
        data = Reference(ws,min_col=2,min_row=max_row2-(max_row2-max_row1)+1,max_row=max_row2,max_col=2)
        title = Reference(ws,min_col=1,min_row=max_row2-(max_row2-max_row1)+1,max_row=max_row2,max_col=1)
        chart2.add_data(data)
        chart2.set_categories(title)
        chart2.height = 10
        chart2.width = 20
        s2 = chart2.series[0]
        s2.dLbls = DataLabelList() 
        s2.dLbls.showVal = True
        s2.dLbls.showPercent = True
        wb.active=1
        ws=wb.active
        ws.add_chart(chart2,'S1')
        


        #电脑配置对比图
        wb.active=2
        ws = wb.active
        sql_pie2 = 'select level,count(deviceid) from d_base where state <> ? group by level'
        result = self.cursor.execute(sql_pie2,'已报废').fetchall()
        for n in result:
            ws.append([n[0],n[1]])
        max_row3 = ws.max_row
        chart3 = PieChart()
        chart3.title = '电脑配置占比图'
        data = Reference(ws,min_col=2,min_row=max_row3-(max_row3-max_row2)+1,max_row=max_row3,max_col=2)
        title = Reference(ws,min_col=1,min_row=max_row3-(max_row3-max_row2)+1,max_row=max_row3,max_col=1)
        chart3.add_data(data)
        chart3.set_categories(title)
        chart3.height = 10
        chart3.width = 20
        s3 = chart3.series[0]
        s3.dLbls = DataLabelList() 
        s3.dLbls.showVal = True
        s3.dLbls.showPercent = True
        #s1.dLbls.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)], bodyPr=rot)

        wb.active=1
        ws=wb.active
        ws.add_chart(chart3,'S23')

    

        wb.active=0
        wb.save('设备维保.xlsx')
        os.startfile('设备维保.xlsx')

    def wb_record_del(self):
        index = self.Tb_WB.currentIndex()
        wb_id = self.Tb_WB.model().index(index.row(),0).data()
        if wb_id is None:
            self.statusBar.showMessage('请先选择维保编号!')
            return
        sql = "delete from wb_record where wb_id = ?"
        counter = self.cursor.execute(sql,wb_id).rowcount
        if counter > 0:
            self.statusBar.showMessage("累计删除"+str(counter)+"条记录!")
        self.cursor.commit()


    def Show_ch_Record(self):
        sql = 'select ch_id,ch_date,user_old,user_new,ch_reason,remark from ch_record where deviceid = ?'
        result = self.cursor.execute(sql,self.pc_id).fetchall()
        model = QStandardItemModel(len(result),6)
        model.setHorizontalHeaderLabels(['变更编号','变更日期','旧使用者','新使用者','变更原因','备注'])
        for row in range(len(result)):
            for column in range(6):
                item = QStandardItem(result[row][column])
                model.setItem(row,column,item)
        self.tableView_Ch.setModel(model)

  

    def Dev_trace(self):
        #设备的追踪信息写入到excel表格中，包括硬件信息，维保记录，变更记录
        #取得硬件设备信息
        sql = 'select deviceid,supplier,dname,ios,level,cpu,mem,solid_disk,harddisk,buydate,ip_addr,position from d_base where deviceid =?'
        result = self.cursor.execute(sql,self.pc_id).fetchone()
        #写入excel表格
        wb=openpyxl.load_workbook('模板.xlsx')
        ws= wb.active
        #作为重新加载的excel表格，需要提前清除掉前面的内容
        ws.cell(row=4,column=8,value=None)
        ws.cell(row=6,column=2,value=None)
        ws.cell(row=6,column=4,value=None)
        ws.cell(row=6,column=6,value=None)
        ws.cell(row=6,column=8,value=None)
        ws.cell(row=7,column=2,value=None)
        ws.cell(row=7,column=4,value=None)
        ws.cell(row=7,column=6,value=None)
        ws.cell(row=7,column=8,value=None)
        ws.cell(row=8,column=2,value=None)
        ws.cell(row=8,column=6,value=None)
        ws.cell(row=8,column=8,value=None)
        #再写入新内容
        ######################################################
        ws.cell(row=4,column=8,value=result[0])
        ws.cell(row=6,column=2,value=result[1])
        ws.cell(row=6,column=4,value=result[2])
        ws.cell(row=6,column=6,value=result[3])
        ws.cell(row=6,column=8,value=result[4])
        ws.cell(row=7,column=2,value=result[5])
        ws.cell(row=7,column=4,value=result[6])
        ws.cell(row=7,column=6,value=result[7])
        ws.cell(row=7,column=8,value=result[8])
        ws.cell(row=8,column=2,value=result[9])
        ws.cell(row=8,column=6,value=result[10])
        ws.cell(row=8,column=8,value=result[11])

        #变更信息
        #清除旧内容
        for row in ws['A13:D19']:
            for cell in row:
                cell.value = None

        sql = 'select ch_date,user_old,user_new,remark from ch_record where deviceid = ?'
        result_ch = self.cursor.execute(sql,self.pc_id).fetchall()
        for n in result_ch:
            ws.append([n[0],n[1],n[2],n[3]])
        #将数据上移29行，若增加了excel的栏位，需要更改相应的行数
        ws.move_range('A'+str(ws.max_row-len(result_ch)+1)+':D'+str(ws.max_row+1),-33,0)

        #维保记录
        #清除旧内容
        for row in ws['A23:D40']:
            for cell in row:
                cell.value = None
        #for row in ws['G23:G35']:
        #    for cell in row:
                cell.value = None
        sql = 'select wb_date,wb_type,operater,wb_desc from wb_record where deviceid = ?'
        result_wb = self.cursor.execute(sql,self.pc_id).fetchall()
        if len(result_wb) > 0:
            for n in result_wb:
                ws.append([n[0],n[1],n[2],n[3]])
            if len(result_wb) ==0:
                ws.move_range('A'+str(ws.max_row-len(result_wb)+1)+':D'+str(ws.max_row+1),-24,0)
            else:
                ws.move_range('A'+str(ws.max_row-len(result_wb)+1)+':D'+str(ws.max_row+1),-23-len(result_ch)-1,0)
            
            #E列为合并并无格，需单独列出移动
            #if len(result_wb) ==0:
            #    ws.move_range('E'+str(ws.max_row-len(result_wb)+1)+':E'+str(ws.max_row+1),-19,2)
            #else:
            #    ws.move_range('E'+str(ws.max_row-len(result_wb)+1)+':E'+str(ws.max_row+1),-18-len(result_ch)-1,2)
        
        #移动后格式感觉会变，所性重新写一下格式
        xborder = Border(left=Side(border_style='thin',color='000000'),right=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
        for row in ws['A12:D19']:
            for cell in row:
                cell.border = xborder
                cell.font = Font(name='微软雅黑',size=11)
        for row in ws['A23:G40']:
            for cell in row:
                cell.border = xborder
                cell.font = Font(name='微软雅黑',size=11)
        
        wb.save('模板.xlsx')
        os.startfile('模板.xlsx')
        


if __name__ =="__main__":
    app = QApplication()
    win = weibao()
    win.show()
    app.exec_()
